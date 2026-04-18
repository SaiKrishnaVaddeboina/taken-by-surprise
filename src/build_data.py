"""
build_data.py
--------------
Data pipeline for the CS 573 Final Project: Surprise Map Web Application.

Joins the FBI 2024 Cargo Theft state-level table with U.S. Census Bureau
2024 incorporated-place population estimates, then computes per-capita
rates and a Bayesian Surprise score using the de Moivre's funnel model
from Correll & Heer (2017) as adapted by Ndlovu, Shrestha, & Harrison
(2023, arXiv:2307.15138).

Inputs
------
  ../../cargo_theft_by_state_2024_data_modified.xlsx
      State-level cargo theft (agencies, incidents, stolen value,
      recovered value, percent recovered). Source: FBI CDE, 2024.

  ../../SUB-IP-EST2024-POP.xlsx
      Annual Population Estimates for Incorporated Places, April 1,
      2020 to July 1, 2024. Source: U.S. Census Bureau, Vintage 2024.

Outputs
-------
  ../data/state_populations_2024.csv
      State-level 2024 population (sum of incorporated places).

  ../data/cargo_theft_surprise_2024.csv
      Fully joined, analysis-ready dataset with computed Bayesian
      Surprise scores per state.

Methodology
-----------
Surprise is operationalised as the two-sided p-value under a de Moivre
funnel model of the per-capita cargo-theft incident rate:

    p_hat = total_incidents / total_population     (national rate)
    Z_i   = (rate_i - p_hat) / sqrt(p_hat*(1-p_hat)/pop_i)
    surprise_p_i = 2 * (1 - Phi(|Z_i|))
    surprise_score_i = -log10(surprise_p_i)

Low surprise_p (equivalently high surprise_score) indicates a state
whose cargo-theft incident rate deviates more than the expected
sampling variation around the national mean, given the state's
population. This is the same statistic visualised by Correll & Heer
(2017, Fig. 4) for unemployment and by Ndlovu et al. (2023) for
Covid-19 vaccination and poverty.

Author: Sai Krishna Vaddeboina, WPI M.S. Fintech (grad May 2026).
Team: McAlarney, Hu, Farquharson, Vaddeboina. CS 573, Spring 2026.
"""

from __future__ import annotations

import csv
import math
import os
import re
from typing import Dict, List, Tuple

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
DATA_OUT = os.path.abspath(os.path.join(HERE, "..", "data"))

CARGO_FILE = os.path.join(ROOT, "cargo_theft_by_state_2024_data_modified.xlsx")
POP_FILE = os.path.join(ROOT, "SUB-IP-EST2024-POP.xlsx")

os.makedirs(DATA_OUT, exist_ok=True)


# --------------------------------------------------------------------------- #
#  Helpers                                                                    #
# --------------------------------------------------------------------------- #

US_STATES: Dict[str, str] = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}

STATE_FIPS: Dict[str, str] = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08",
    "CT": "09", "DE": "10", "DC": "11", "FL": "12", "GA": "13", "HI": "15",
    "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21",
    "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39",
    "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46",
    "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56",
}


def normal_cdf(x: float) -> float:
    """Standard-normal CDF via math.erf (no SciPy dependency)."""
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


# --------------------------------------------------------------------------- #
#  1. Aggregate incorporated-place populations up to the state level.         #
# --------------------------------------------------------------------------- #

def load_state_populations(path: str) -> Dict[str, int]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    by_state: Dict[str, int] = {name: 0 for name in US_STATES}

    # Rows 1-4 are headers; data starts row 5. Column A is "Place, State",
    # column G is the 2024 estimate (index 6).
    pattern = re.compile(r",\s*([A-Z][A-Za-z ]+)\s*$")
    for row in ws.iter_rows(min_row=5, values_only=True):
        label = row[0]
        pop_2024 = row[6]
        if not label or pop_2024 is None:
            continue
        m = pattern.search(str(label))
        if not m:
            continue
        state = m.group(1).strip()
        # Fix common edge cases.
        state = state.replace("District of Columbia", "District of Columbia")
        if state in by_state and isinstance(pop_2024, (int, float)):
            by_state[state] += int(pop_2024)
    return by_state


# --------------------------------------------------------------------------- #
#  2. Load the state-level cargo-theft table.                                 #
# --------------------------------------------------------------------------- #

def load_cargo_theft(path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        state, agencies, incidents, stolen, recovered, pct = row[:6]
        if not state or state == "State":
            continue
        try:
            rows.append({
                "state_name": str(state).strip(),
                "agencies_reporting": int(agencies or 0),
                "incidents": int(incidents or 0),
                "stolen_value_usd": int(stolen or 0),
                "recovered_value_usd": int(recovered or 0) if isinstance(recovered, (int, float)) else 0,
                "percent_recovered": float(pct) if isinstance(pct, (int, float)) else None,
            })
        except (TypeError, ValueError):
            continue
    return rows


# --------------------------------------------------------------------------- #
#  3. Join + compute Bayesian Surprise scores.                                #
# --------------------------------------------------------------------------- #

def compute_surprise(theft_rows: List[Dict], pops: Dict[str, int]) -> List[Dict]:
    """
    Bayesian Surprise via the de Moivre funnel. Computes three complementary
    surprise metrics to handle the challenge that cargo-theft incidents are
    rare events (~5 per 100k nationally), which causes classical p-value
    saturation in the funnel model:

      1. z_score          - raw signed z, interpretable but very long-tailed.
      2. surprise_p       - two-sided p-value; saturates at 0 for most states.
      3. surprise_score   - -log10(p), also saturates.
      4. surprise_display - sign(z) * sqrt(|z|); bounded non-saturating
                             divergent metric used for color encoding.
      5. value_surprise   - analog of (4) built on stolen-value-per-capita.
    """
    reporting_states = {r["state_name"] for r in theft_rows}
    total_incidents = sum(r["incidents"] for r in theft_rows)
    total_population = sum(p for s, p in pops.items() if p > 0 and s in reporting_states)
    p_hat = total_incidents / total_population if total_population else 0.0

    # Value-weighted global baseline ($ stolen per person nationally).
    total_value = sum(r["stolen_value_usd"] for r in theft_rows)
    v_hat = total_value / total_population if total_population else 0.0
    # Cross-state variance of value-per-capita (for a Gaussian surprise model
    # that doesn't assume the rare-event proportion structure).
    vpc = []
    for r in theft_rows:
        pop = pops.get(r["state_name"], 0)
        if pop > 0:
            vpc.append(r["stolen_value_usd"] / pop)
    v_mean = sum(vpc) / len(vpc) if vpc else 0.0
    v_var = sum((x - v_mean) ** 2 for x in vpc) / len(vpc) if vpc else 0.0
    v_sd = math.sqrt(v_var) if v_var > 0 else 1.0

    enriched: List[Dict] = []
    for r in theft_rows:
        state = r["state_name"]
        pop = pops.get(state, 0)
        if pop <= 0:
            continue
        k = r["incidents"]
        rate = k / pop
        value_per_capita = r["stolen_value_usd"] / pop if pop else 0.0

        # --- Incident-rate surprise (de Moivre funnel) -------------------- #
        se = math.sqrt(p_hat * (1.0 - p_hat) / pop) if pop else 0.0
        z = (rate - p_hat) / se if se > 0 else 0.0
        surprise_p = 2.0 * (1.0 - normal_cdf(abs(z)))
        surprise_p_clipped = max(surprise_p, 1e-30)
        surprise_score = -math.log10(surprise_p_clipped)
        surprise_display = math.copysign(math.sqrt(abs(z)), z) if z != 0 else 0.0

        # --- Value-per-capita surprise (Gaussian across states) ----------- #
        z_value = (value_per_capita - v_mean) / v_sd if v_sd > 0 else 0.0
        value_surprise = math.copysign(math.sqrt(abs(z_value)), z_value) if z_value != 0 else 0.0

        enriched.append({
            **r,
            "state_abbr": US_STATES.get(state, ""),
            "state_fips": STATE_FIPS.get(US_STATES.get(state, ""), ""),
            "population_2024": pop,
            "incidents_per_100k": round(rate * 100_000, 3),
            "stolen_value_per_capita_usd": round(value_per_capita, 2),
            "national_rate_per_100k": round(p_hat * 100_000, 4),
            "national_value_per_capita_usd": round(v_hat, 2),
            "expected_incidents": round(p_hat * pop, 1),
            "z_score": round(z, 3),
            "surprise_p": surprise_p,
            "surprise_score": round(surprise_score, 3),
            "surprise_display": round(surprise_display, 3),
            "z_value": round(z_value, 3),
            "value_surprise": round(value_surprise, 3),
            "direction": "higher_than_expected" if z > 0 else ("lower_than_expected" if z < 0 else "at_expected"),
        })
    return enriched


# --------------------------------------------------------------------------- #
#  4. CSV writers.                                                            #
# --------------------------------------------------------------------------- #

def write_csv(rows: List[Dict], path: str, fields: List[str]) -> None:
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fields})


def main() -> None:
    print("[1/3] Aggregating state populations from", os.path.basename(POP_FILE))
    pops = load_state_populations(POP_FILE)
    nonzero = {s: p for s, p in pops.items() if p > 0}
    print(f"      -> {len(nonzero)}/51 states with population coverage")

    print("[2/3] Loading cargo theft table from", os.path.basename(CARGO_FILE))
    theft = load_cargo_theft(CARGO_FILE)
    print(f"      -> {len(theft)} state rows")

    print("[3/3] Computing Bayesian Surprise (de Moivre funnel)")
    enriched = compute_surprise(theft, pops)
    print(f"      -> {len(enriched)} state rows with surprise scores")

    pop_fields = ["state_name", "state_abbr", "state_fips", "population_2024"]
    pop_rows = [
        {
            "state_name": s,
            "state_abbr": US_STATES[s],
            "state_fips": STATE_FIPS.get(US_STATES[s], ""),
            "population_2024": p,
        }
        for s, p in sorted(pops.items()) if p > 0
    ]
    write_csv(pop_rows, os.path.join(DATA_OUT, "state_populations_2024.csv"), pop_fields)

    cargo_fields = [
        "state_name", "state_abbr", "state_fips",
        "agencies_reporting", "incidents", "stolen_value_usd",
        "recovered_value_usd", "percent_recovered",
        "population_2024",
        "national_rate_per_100k", "national_value_per_capita_usd",
        "incidents_per_100k", "expected_incidents",
        "stolen_value_per_capita_usd",
        "z_score", "surprise_p", "surprise_score", "surprise_display",
        "z_value", "value_surprise", "direction",
    ]
    write_csv(enriched, os.path.join(DATA_OUT, "cargo_theft_surprise_2024.csv"), cargo_fields)

    # Quick sanity printout for the top 10 most surprising states.
    print("\nTop 10 states by |signed-surprise| on per-capita cargo-theft incidents:")
    for r in sorted(enriched, key=lambda x: -abs(x["surprise_display"]))[:10]:
        print(f"  {r['state_abbr']}  z={r['z_score']:+9.2f}  "
              f"rate/100k={r['incidents_per_100k']:>8.2f}  "
              f"display={r['surprise_display']:>+6.2f}  "
              f"({r['direction']})")

    print("\nTop 10 states by |value-surprise| (stolen $/capita deviation):")
    for r in sorted(enriched, key=lambda x: -abs(x["value_surprise"]))[:10]:
        print(f"  {r['state_abbr']}  z_value={r['z_value']:+6.2f}  "
              f"$/capita={r['stolen_value_per_capita_usd']:>8.2f}  "
              f"display={r['value_surprise']:>+6.2f}")


if __name__ == "__main__":
    main()
